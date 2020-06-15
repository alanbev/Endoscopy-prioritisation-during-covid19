import datetime
import openpyxl
import sys

#New in this version:

#Program creates template spreadsheet with headers read from patient data object- template changes if object edited

#Program reads spreadsheets in this format to create list of patient dictionary objects

#Data from spreadsheet is cleaned. y, yes, t and true (not case-sensitive) are interpreted as True in boolean fields.
#Missing booleans are interpreted as false.  Option given to input missing non-booleans manually.  Entries with non-numeric hospital
#number are deleted with a warning.

#Program writes new Excel workbook with cleaned data on first sheet and risk adjusts on second sheet.
#Option to display data on screen before writing to workbook.

#Dataset now includes 2 values for FIT to allow for double test. Maximum value of FIT is used to assess risk.

#Patients with both FIT values <10 no longer have risk augmented for time on waiting list.

#database write NOT implemented in this version.
#preview option is very rough - needs formatting work

def set_risk_adjusts():
    """creates dictionary object containing risk adjustments 1-6 are base risks as per protocol, age adjust_per decade adds risk per decade from 50 to 80, fem sex removes risk from females over 60. prior ix scales risk if ct in last year or colonoscopy in last 3 years, time_inc adds risk weighting for each week on waiting list"""
    risk_adjust = {
        1: 20,
        2: 16,
        3: 12,
        4: 8,
        5: 4,
        6: 0,
        "age_adjust_per_decade": 1,
        "fem_sex": -1,       #accounts for lag in female age / incidence curve compared to male
        "prior_ix": 0.75,    #currently does not distinguish between Ix on this episode and either CT<1yr or colonoscopy<3yr
        "time_inc": 1        #change if slower or faster wait-list advancement required.  No advancement if FIT <10.
    }
    return risk_adjust


def make_patient():
    """Creates template patient dictionary object.  If fields are added or removed the headers on new template spreadshees will be changed accordingly"""
    patient = {
        "hosp_no": 0,
        "forename": "Joe",
        "surname": "Bloggs",
        "dob": datetime.date(1950, 7, 7),
        "female": True,
        "palp_mass": True,
        "ct_mass": True,
        "bleeding": True,
        "loose_stools": True,
        "fit_done": True,
        "fit_value1": 0,
        "fit_value2": 0,
        "prior_colonoscopy": True,
        "prior_CT": True,
        "date_listed": datetime.datetime(2020, 3, 25),
        "ct_pending":True,
    }
    return patient

def make_blank_spreadsheet(example):
    """makes spreadsheet with headers matching keys of patient dictionary"""
    wb = openpyxl.Workbook()
    ws = wb.active
    write_column = 1
    for key in example:
        ws.cell(row=1, column=write_column).value = key
        write_column += 1
    wb.save(filename="blank patients.xlsx")
    print("A new spreadsheet has been created called blank patients- you should rename this before use to prevent over-writing")

def read_spreadsheet(empty_dict):
    """opens excel workbook and reads contents into dictionary"""
    wb_to_use = input("Please enter the name of the sheet you want to use (without the .xlsx file extension)")

    try:
        wb = openpyxl.load_workbook(wb_to_use + ".xlsx")
    except:
        try_again = ""
        print("The spreadsheet you requested could not be opened. It may be open in another application or the directory path may be incorrect.")
        while try_again == "":
            try_again = input("Do you want to try again (Y/N?)")
            if try_again.upper() == "Y":
                wb_to_use = ""
                retry = read_spreadsheet(empty_dict)
                return retry
            else:
                print ("OK-Program will terminate")
                sys.exit()

    ws = wb.active
    entries = ws.max_row #finds end of entries on worksheet
    all_patients = [] #creates list to store all patient dictionary items in
    for record in range(2, entries):
        current_patient = dict(empty_dict)
        data_item = 1

        for key in current_patient:
            data = ws.cell(row=record, column=data_item).value
            current_patient[key] = data
            data_item += 1
        all_patients.append(current_patient)
    return (all_patients, wb_to_use)


def clean_data(all_patients_list, patient):
    """modifies input from spreadsheet to remove empty rows and create valid Booleans.  Missing boolean data is assumed to be False"""
    index = 0
    for dictionary in all_patients_list:
        if not str(dictionary["hosp_no"]).isnumeric():
            all_patients_list.pop(index)
            print("Entry from row %i with patient name %s, %s, removed due to missing or invalid hospital number./n If this is an error, please check the spreadsheet and then re-import it ",(index + 2, dictionary["forename"],dictionary["surname"]))
        else:
            for key in patient:

                if isinstance(patient[key], bool):
                    if dictionary[key] is None:
                        dictionary[key] = "False"

                    if dictionary[key].upper() in ["T", "Y"]:
                        dictionary[key] = True
                    else:
                        dictionary[key] = False
                else:
                    if dictionary[key] is None:
                        message_for_input = " Missing data for {0} in row {1}, hospital number {2}, \n Enter this now or type x to ignore \n This may lead to the program terminating- if this occurs fix the spreadsheet and re-run".format(key, index+1, dictionary["hosp_no"])
                        manual_input=input(message_for_input)
                        if manual_input.lower() != "x":
                            dictionary[key] = manual_input
                index += 1
    return all_patients_list


def symptom_risk(patient, risk_adjust):
    """calculates and returns baseline risk from symptoms and FIT result using risk weightings specified in risk_adjust dictionary """

    fit_value = max(int(patient["fit_value1"]),int(patient["fit_value2"]))
    if patient["palp_mass"] or patient["ct_mass"]:
        return risk_adjust[1]
    if (patient["loose_stools"] and patient["bleeding"]) or fit_value > 100:
        return risk_adjust[2]
    if patient["bleeding"]:
        return risk_adjust[3]
    if patient["loose_stools"] and fit_value > 10:
        return risk_adjust[4]
    if fit_value > 10:
        return risk_adjust[5]

    return risk_adjust[6]


def age_sex_risk_adjust(patient, risk_adjust):
    """adjusts risk up for age over 50 and down for female according to risk adjust settings - returns adjustment factor to be applied to patient risk outside the function """
    this_day = datetime.date.today()
    patient_dob = (patient["dob"]).date()
    age_object = this_day - patient_dob
    age = age_object.days / 365
    if age < 50:
        unisex_risk = 0
    elif age > 80:
        unisex_risk = 4 * risk_adjust["age_adjust_per_decade"]
    else:
        unisex_risk = int((age - 50) / 10 * risk_adjust["age_adjust_per_decade"])
    if age > 60 and patient["female"]:
        return unisex_risk - risk_adjust["fem_sex"]
    else:
        return unisex_risk


def prior_investigation_risk_adjust(patient, risk_adjust):
    """scales risk down if previous clear ct in last year or clear colonoscopy in last 3 years"""
    if patient["prior_CT"] or patient["prior_colonoscopy"]:
        return patient["risk2"] * risk_adjust["prior_ix"]
    else:
        return patient["risk2"]


def waiting_time_at_entry_adjust(patient, risk_adjust):
    """calculates and returns increment for weeks on waiting list"""
    date_listed = patient["date_listed"].date()
    date_entered = datetime.date.today()
    wait_object = date_entered - date_listed
    weeks_waiting_at_entry = int(wait_object.days / 7)
    return (weeks_waiting_at_entry * risk_adjust["time_inc"])

def write_to_db(patient):
    """writes inputted data to database"""
    pass
    #tables for 1. demographics, 2. symptoms and prior ix, 3 FIT status
    #?do in SQLLITE or ?via SQLALCHEMY

def find_needing_fit():
    """runs dB query to find patients in DB FIT table with no FIT test and no documented bleeding and outputs list"""
    pass

def enter_fit_result():
    """Allows entry of FIT test result and re-runs risk assessment"""
    pass



def risk_adjustment_on_entry(cleaned_list, risk_adjust):
    """main control function for adjusting patient risk on entry"""
    for each_patient in cleaned_list:
        risk1 = symptom_risk(each_patient, risk_adjust)
        each_patient["risk1"] = risk1
        risk2 = age_sex_risk_adjust(each_patient, risk_adjust)
        each_patient["risk2"] = each_patient["risk1"] + risk2
        risk3 = prior_investigation_risk_adjust(each_patient, risk_adjust)
        each_patient["risk3"] = risk3
        wait_time_adjust = waiting_time_at_entry_adjust(each_patient, risk_adjust)
        if each_patient["risk1"] == risk_adjust[6]:
            each_patient["risk4"] = each_patient["risk3"]#blocks waiting list advancement for FIT -ve patients
        else:
            each_patient["risk4"] = each_patient["risk3"] + wait_time_adjust

    return cleaned_list



def add_comments(risk_adjusted_patients):
    """Adds comments regarding status of data items to patient dictionary object"""
    for each_patient in risk_adjusted_patients:
        if each_patient["fit_done"]:
            each_patient["comment"] = "Risk includes FIT test"
        elif each_patient["bleeding"]:
            each_patient["comment"]= "FIT test not required"
        else:
            each_patient["comment"] = "Provisional Risk Pending FIT test"
        if each_patient["ct_pending"]:
            each_patient["comment"] = each_patient["comment"] + "-(CT currently arranged -result pending)"
    return risk_adjusted_patients


def write_spreadsheet(output_list, file_name, cleaned_list):
    """writes risk calculations to new spreadsheet """
    write_file_name = file_name + " risk statification.xlsx"
    write_book = openpyxl.Workbook()
    write_sheet1= write_book.create_sheet("cleaned data", 0)
    headers = list(cleaned_list[0].keys())
    write_sheet1 = write_data_to_sheet(write_sheet1, cleaned_list, headers, headers)

    headers = (
        "hosp_no",
        "forename",
        "surname",
        "dob",
        "FIT & Symptom risk",
        "Age Adjusted",
        "Prior Ix Adjusted",
        "Delay Adjusted",
        "Comments"
    )
    data_items = (
        "hosp_no",
        "forename",
        "surname",
        "dob",
        "risk1",
        "risk2",
        "risk3",
        "risk4",
        "comment"
    )
    write_sheet2 = write_book.create_sheet("risks for prioritisation",1)
    write_sheet2 = write_data_to_sheet(write_sheet2,output_list, headers, data_items )

    try:
        write_book.save(write_file_name)
    except:
        write_book.save("copy of" + write_file_name)


def write_data_to_sheet(current_write_sheet, current_output_list, headers, data_items):
    """called by write_spreadsheet to convert dictionary to worksheet"""
    write_column = 1
    for item in headers:
        current_write_sheet.cell(row=1, column=write_column).value = item
        write_column += 1

    write_row = 2
    write_column = 1
    for dictionary in current_output_list:
        for item in data_items:
            if isinstance(item, datetime.date):
                current_write_sheet.cell(row=write_row, column=write_column).style.number_format.format_code = "DD/MM/YYYY"
            current_write_sheet.cell(row=write_row, column=write_column).value = dictionary[item]
            write_column += 1
        write_column = 1
        write_row += 1
    return current_write_sheet


def preview_data(commented_list, spreadsheet_name, cleaned_list):
    """displays risk adjustments to screen"""
    print("hosp_no forename surname    FIT & Symptom risk  Age Adjusted  Prior Ix Adjusted   Delay Adjusted  Comments")
    for dictionary in commented_list:
        print("{0}       {1}         {2}               {3}            {4}               {5}                {6}         {7}".format(dictionary["hosp_no"],dictionary["forename"],dictionary["surname"],dictionary["risk1"],dictionary["risk2"],dictionary["risk3"],dictionary["risk4"],dictionary["comment"]))
    output_option = ""
    while not output_option in ("y","n"):
        output_option = input("Do you want to write these results to a spreadsheet now?").lower()
        if output_option[0] == "y":
            write_spreadsheet(commented_list, spreadsheet_name, cleaned_list)
        else:
            output_option=""
            while not output_option[0] in ("y","n"):
                output_option = input("Do you want to exit without saving results? (y, n").lower()
            if output_option[0] == "n":
                preview_data(commented_list, spreadsheet_name, cleaned_list)
            else:
                sys.exit()

#main
risk_adjust = set_risk_adjusts()
patient = make_patient()
opt_read_wb = ""
while opt_read_wb == "":
    opt_read_wb = input("Do you want to read a spreadsheet?")
if opt_read_wb[0].lower() == "y":
    list_all_patients, spreadsheet_name = read_spreadsheet(patient)
    cleaned_list = clean_data(list_all_patients, patient)
    risk_adjusted_patients = risk_adjustment_on_entry(cleaned_list, risk_adjust)
    commented_list = add_comments(risk_adjusted_patients)
    output_option = ""
    while not output_option in ("w", "p"):
        output_option = input("Write results to excel spreadsheet or Preview? (enter w or p)")
        if output_option == "w":
            write_spreadsheet(commented_list, spreadsheet_name, cleaned_list)
        else:
            preview_data(commented_list, spreadsheet_name, cleaned_list)
opt_make_new_wb = ""
while opt_make_new_wb not in ("y","n"):
    opt_make_new_wb = input("Do you want to create a new spreadsheet?").lower()
if opt_make_new_wb[0].lower() == "y":
    make_blank_spreadsheet(patient)
else:
    sys.exit()






